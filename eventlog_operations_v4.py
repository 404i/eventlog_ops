# Requirements:

# Install tqdm via pip
# Install chainsaw, apt-hunter and hayabusa via git in $USER/git/
# Depending on your Hayabusa version you might need to change the path/binary name in the hayabusa executions snippet.


import argparse
import configparser
import csv
import json
import logging
import os
import shlex
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import requests
except ImportError:  # pragma: no cover - optional dependency
    requests = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


COLOR_RESET = "\033[0m"
COLOR_GREEN = "\033[92m"
COLOR_YELLOW = "\033[93m"
COLOR_BLUE = "\033[94m"
COLOR_RED = "\033[91m"
COLOR_ORANGE = "\033[38;5;214m"
USE_COLOR = sys.stdout.isatty()


def colorize(text: str, color: str) -> str:
    if not USE_COLOR:
        return text
    return f"{color}{text}{COLOR_RESET}"


ARTIFACT_CACHE: Dict[str, Tuple[Tuple, Dict[str, object]]] = {}


def colorize_confidence(value: str) -> str:
    lowered = value.lower()
    if "high" in lowered:
        return colorize(value, COLOR_RED)
    if "medium" in lowered:
        return colorize(value, COLOR_ORANGE)
    if "low" in lowered:
        return colorize(value, COLOR_GREEN)
    return value

# from tqdm import tqdm

# Locate configuration alongside script or executable
try:
    CONFIG_ROOT = Path(__file__).resolve().parent
except NameError:
    CONFIG_ROOT = Path(sys.argv[0]).resolve().parent

CONFIG_FILE = CONFIG_ROOT / "eventlog_tools.ini"

DEFAULT_LLM_SETTINGS = {
    "enabled": "false",
    "endpoint": "http://localhost:1234/v1/chat/completions",
    "model": "openai/gpt-oss-120b",
    "system_prompt": (
        "You are an incident-response assistant. Summarize findings with a concise markdown table "
        "(columns: Tool, Key Finding, Supporting Evidence) followed by a short narrative summary."
    ),
    "temperature": "0.2",
    "max_tokens": "-1",
    "timeout_seconds": "180",
}


def save_config(config: configparser.ConfigParser) -> None:
    with CONFIG_FILE.open("w", encoding="utf-8") as config_file:
        config.write(config_file)


@dataclass
class ToolRun:
    name: str
    command: str
    log_path: Path
    report_path: Path
    extra_reports: List[Path] = field(default_factory=list)
    process: Optional[subprocess.Popen] = field(default=None, init=False)
    return_code: Optional[int] = field(default=None, init=False)


def prompt_for_path(prompt_text: str, default_path: Path) -> Path:
    """Interactively prompt user for a filesystem path, with validation."""

    while True:
        user_input = input(f"{prompt_text} [{default_path}]: ").strip()
        candidate = Path(user_input or default_path).expanduser()

        if candidate.exists():
            return candidate

        retry = input(
            f"Path '{candidate}' does not exist. Press Enter to retry or type 'yes' to accept anyway: "
        ).strip().lower()
        if retry in {"y", "yes"}:
            return candidate


def prompt_yes_no(message: str, default: bool = False) -> bool:
    default_text = "Y/n" if default else "y/N"
    response = input(f"{message} [{default_text}]: ").strip().lower()
    if not response:
        return default
    return response in {"y", "yes"}


def prompt_with_default(message: str, default: str) -> str:
    response = input(f"{message} [{default}]: ").strip()
    return response or default


def collect_configuration() -> configparser.ConfigParser:
    """Run interactive setup for tool paths and LLM integration."""

    print("Initial setup: specify where each tool is installed (absolute paths recommended).")
    git_root = Path.home() / "git"
    hayabusa_default = git_root / "hayabusa" / "hayabusa"
    chainsaw_default = git_root / "chainsaw" / "target" / "release" / "chainsaw"
    sigma_default = git_root / "sigma"
    mapping_default = git_root / "chainsaw" / "mappings" / "sigma-event-logs-all.yml"
    chainsaw_rules_default = git_root / "chainsaw" / "rules"

    apt_hunter_roots = [
        git_root / "APT-Hunter",
        git_root / "APT-Hunter-main",
    ]
    apt_hunter_root = next((root for root in apt_hunter_roots if root.exists()), apt_hunter_roots[0])

    defaults = {
        "hayabusa": hayabusa_default,
        "apt_hunter": apt_hunter_root / "APT-Hunter.py",
        "apt_hunter_python": apt_hunter_root / ".venv" / "bin" / "python",
        "chainsaw": chainsaw_default,
        "sigma": sigma_default,
        "mapping": mapping_default,
        "chainsaw_rules": chainsaw_rules_default,
    }

    paths = {}
    paths["hayabusa"] = prompt_for_path("Path to hayabusa binary", defaults["hayabusa"])
    paths["apt_hunter"] = prompt_for_path("Path to APT-Hunter.py", defaults["apt_hunter"])
    paths["apt_hunter_python"] = prompt_for_path(
        "Python interpreter to run APT-Hunter", defaults["apt_hunter_python"]
    )
    paths["chainsaw"] = prompt_for_path("Path to chainsaw binary", defaults["chainsaw"])
    paths["sigma"] = prompt_for_path("Path to Sigma rules directory", defaults["sigma"])
    paths["mapping"] = prompt_for_path(
        "Path to chainsaw Sigma mapping file", defaults["mapping"]
    )
    paths["chainsaw_rules"] = prompt_for_path(
        "Path to chainsaw rules directory", defaults["chainsaw_rules"]
    )

    print("\nLLM summary configuration (optional - used with --llm-summary).")
    llm_settings: Dict[str, str] = DEFAULT_LLM_SETTINGS.copy()
    if prompt_yes_no("Enable LLM summarization integration?", False):
        llm_settings["enabled"] = "true"
        llm_settings["endpoint"] = prompt_with_default(
            "LLM endpoint URL", llm_settings["endpoint"]
        )
        llm_settings["model"] = prompt_with_default("LLM model identifier", llm_settings["model"])
        llm_settings["system_prompt"] = prompt_with_default(
            "System prompt", llm_settings["system_prompt"]
        )
        llm_settings["temperature"] = prompt_with_default(
            "Sampling temperature", llm_settings["temperature"]
        )
        llm_settings["max_tokens"] = prompt_with_default(
            "Max tokens (-1 for model default)", llm_settings["max_tokens"]
        )
        llm_settings["timeout_seconds"] = prompt_with_default(
            "HTTP timeout in seconds", llm_settings["timeout_seconds"]
        )
    else:
        llm_settings["enabled"] = "false"

    config = configparser.ConfigParser()
    config["tools"] = {key: str(value) for key, value in paths.items()}
    config["llm"] = llm_settings
    save_config(config)

    print(f"Configuration saved to {CONFIG_FILE}. Delete this file to re-run setup.")
    return config


def load_configuration() -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    if not CONFIG_FILE.exists():
        return collect_configuration()

    config.read(CONFIG_FILE, encoding="utf-8")

    required_tool_keys = {
        "hayabusa",
        "apt_hunter",
        "apt_hunter_python",
        "chainsaw",
        "sigma",
        "mapping",
        "chainsaw_rules",
    }

    if "tools" not in config:
        print("Configuration missing 'tools' section. Re-running setup.")
        return collect_configuration()

    tools_section = config["tools"]
    missing = required_tool_keys - set(tools_section.keys())
    if missing:
        print(f"Configuration missing keys: {', '.join(sorted(missing))}. Re-running setup.")
        return collect_configuration()

    if "llm" not in config:
        config["llm"] = DEFAULT_LLM_SETTINGS.copy()
        save_config(config)
    else:
        updated = False
        for key, value in DEFAULT_LLM_SETTINGS.items():
            if key not in config["llm"]:
                config["llm"][key] = value
                updated = True
        if updated:
            save_config(config)

    return config


def get_tool_paths(config: configparser.ConfigParser) -> dict:
    tools_section = config["tools"]
    return {
        key: Path(tools_section[key]).expanduser()
        for key in tools_section
    }


def load_llm_settings(config: configparser.ConfigParser) -> Dict[str, object]:
    llm_section = config["llm"]
    settings = DEFAULT_LLM_SETTINGS.copy()
    settings.update(llm_section)

    def parse_bool(value: str) -> bool:
        return value.strip().lower() in {"1", "true", "yes", "y"}

    def parse_float(value: str, default: float) -> float:
        try:
            return float(value)
        except ValueError:
            return default

    def parse_int(value: str, default: int) -> int:
        try:
            return int(value)
        except ValueError:
            return default

    parsed = {
        "enabled": parse_bool(settings.get("enabled", "false")),
        "endpoint": settings.get("endpoint", DEFAULT_LLM_SETTINGS["endpoint"]),
        "model": settings.get("model", DEFAULT_LLM_SETTINGS["model"]),
        "system_prompt": settings.get("system_prompt", DEFAULT_LLM_SETTINGS["system_prompt"]),
        "temperature": parse_float(settings.get("temperature", "0.2"), 0.2),
        "max_tokens": parse_int(settings.get("max_tokens", "-1"), -1),
        "timeout_seconds": parse_int(settings.get("timeout_seconds", "180"), 180),
    }
    return parsed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run Hayabusa, APT-Hunter, and Chainsaw against a log directory."
    )
    parser.add_argument(
        "target",
        nargs="?",
        help="Path to the directory containing event logs (defaults to an interactive prompt)",
    )
    parser.add_argument(
        "--archive",
        action="store_true",
        help="Zip reports and logs into the target directory when processing completes.",
    )
    parser.add_argument(
        "--retry-failed",
        action="store_true",
        help="Automatically retry any tools that exit with a non-zero status.",
    )
    parser.add_argument(
        "--update-tools",
        action="store_true",
        help="Update Hayabusa, APT-Hunter, Chainsaw, and Sigma repositories before execution.",
    )
    parser.add_argument(
        "--llm-summary",
        action="store_true",
        help="Send report outputs to the configured LLM endpoint for a consolidated summary.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable verbose debug logging to the console.",
    )
    parser.add_argument(
        "--debug-log",
        type=Path,
        help="Optional path to write detailed debug logs.",
    )
    return parser.parse_args()


def resolve_target_directory(args: argparse.Namespace) -> Path:
    if args.target:
        candidate = Path(args.target).expanduser()
    else:
        while True:
            user_input = input("Path to event log directory: ").strip()
            candidate = Path(user_input).expanduser()
            if candidate.exists():
                break
            print(f"Provided path '{candidate}' does not exist. Please try again.")
    if not candidate.exists():
        raise FileNotFoundError(f"Target directory '{candidate}' does not exist.")
    if not candidate.is_dir():
        raise NotADirectoryError(f"Target path '{candidate}' is not a directory.")
    return candidate.resolve()


def human_size(num_bytes: int) -> str:
    if num_bytes < 1024:
        return f"{num_bytes} B"
    for unit in ["KB", "MB", "GB", "TB"]:
        num_bytes /= 1024.0
        if num_bytes < 1024:
            return f"{num_bytes:.1f} {unit}"
    return f"{num_bytes:.1f} PB"


def artifact_size_bytes(path: Path) -> int:
    try:
        if path.is_dir():
            total = 0
            for child in path.rglob("*"):
                if child.is_file():
                    total += child.stat().st_size
            return total
        if path.is_file():
            return path.stat().st_size
    except OSError:
        pass
    return 0


def artifact_signature(path: Path) -> Tuple:
    try:
        stat = path.stat()
        return (
            True,
            path.is_dir(),
            getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1e9)),
            stat.st_size,
        )
    except FileNotFoundError:
        return (False, False, None, None)


def analyze_csv_artifact(path: Path, max_preview: int = 3) -> Dict[str, object]:
    result: Dict[str, object] = {
        "type": "csv",
        "path": str(path),
        "size": human_size(path.stat().st_size),
        "rows": 0,
        "columns": [],
        "preview": [],
    }
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            try:
                header = next(reader)
            except StopIteration:
                header = []
            result["columns"] = header
            for idx, row in enumerate(reader, start=1):
                if idx <= max_preview:
                    result["preview"].append(row[:6])
                result["rows"] = idx
    except OSError as exc:
        result["error"] = f"Failed to read CSV: {exc}"
    return result


def analyze_xlsx_artifact(path: Path, max_sheets: int = 4, max_preview: int = 3) -> Dict[str, object]:
    result: Dict[str, object] = {
        "type": "xlsx",
        "path": str(path),
        "size": human_size(path.stat().st_size),
        "sheet_count": 0,
        "sheets": [],
    }

    if load_workbook is None:
        result["error"] = "openpyxl not installed"
        return result

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - depends on external file
        result["error"] = f"Failed to open workbook: {exc}"
        return result

    try:
        sheet_names = wb.sheetnames
        result["sheet_count"] = len(sheet_names)
        for sheet_index, sheet_name in enumerate(sheet_names[:max_sheets]):
            ws = wb[sheet_name]
            sheet_info: Dict[str, object] = {
                "name": sheet_name,
                "rows": ws.max_row or 0,
                "columns": ws.max_column or 0,
                "preview": [],
            }
            try:
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_index > max_preview:
                        break
                    truncated = []
                    for cell in row[:6]:
                        if cell is None:
                            truncated.append("")
                        else:
                            text = str(cell)
                            truncated.append(text[:80])
                    sheet_info["preview"].append(truncated)
            except Exception as exc:  # pragma: no cover
                sheet_info["preview_error"] = f"Failed to iterate rows: {exc}"
            result["sheets"].append(sheet_info)
    finally:
        wb.close()

    return result


def _describe_artifact_uncached(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {
            "path": str(path),
            "exists": False,
            "type": None,
            "size_bytes": 0,
        }
    try:
        if path.is_dir():
            children = []
            for child in sorted(path.iterdir())[:20]:
                child_summary = describe_artifact(child)
                entry = {
                    "name": child.name,
                    "type": child_summary.get("type"),
                    "size": child_summary.get("size"),
                    "size_bytes": child_summary.get("size_bytes", 0),
                }
                details = child_summary.get("details", {})
                if child_summary.get("type") == "csv":
                    entry["rows"] = details.get("rows")
                    entry["columns"] = details.get("columns")
                elif child_summary.get("type") in {"xlsx", "xls", "xlsm"}:
                    entry["sheet_count"] = child_summary.get("sheet_count")
                if child_summary.get("error"):
                    entry["error"] = child_summary.get("error")
                children.append(entry)
            total_bytes = artifact_size_bytes(path)
            return {
                "type": "directory",
                "path": str(path),
                "size": human_size(total_bytes),
                "size_bytes": total_bytes,
                "exists": True,
                "contents": children,
            }
    except OSError as exc:
        return {
            "type": "directory",
            "path": str(path),
            "exists": True,
            "size_bytes": 0,
            "error": f"Failed to enumerate directory: {exc}",
        }

    lower = path.suffix.lower()
    if lower == ".csv":
        csv_info = analyze_csv_artifact(path)
        csv_info["exists"] = True
        csv_info["size_bytes"] = path.stat().st_size
        csv_info["details"] = {
            "rows": csv_info.get("rows", 0),
            "columns": csv_info.get("columns", []),
            "preview": csv_info.get("preview", []),
        }
        return csv_info
    if lower in {".xlsx", ".xlsm", ".xls"}:
        xlsx_info = analyze_xlsx_artifact(path)
        xlsx_info["exists"] = True
        xlsx_info["size_bytes"] = path.stat().st_size
        xlsx_info["details"] = {
            "sheets": xlsx_info.get("sheets", []),
        }
        return xlsx_info
    return {
        "type": lower.lstrip("."),
        "path": str(path),
        "size": human_size(path.stat().st_size),
        "size_bytes": path.stat().st_size,
        "exists": True,
    }


def describe_artifact(path: Path) -> Dict[str, object]:
    cache_key = str(path)
    signature = artifact_signature(path)
    cached = ARTIFACT_CACHE.get(cache_key)
    if cached and cached[0] == signature:
        return cached[1]

    summary = _describe_artifact_uncached(path)
    ARTIFACT_CACHE[cache_key] = (signature, summary)
    return summary


def get_artifact_summaries(tool: ToolRun) -> List[Dict[str, object]]:
    summaries: List[Dict[str, object]] = []
    seen = set()
    for path in [tool.report_path, *tool.extra_reports]:
        key = str(path.resolve())
        if key in seen:
            continue
        seen.add(key)
        summaries.append(describe_artifact(path))
    return summaries


def prepare_tools(
    target_dir: Path,
    folder_name: str,
    report_folder: Path,
    log_folder: Path,
    tool_paths: dict,
) -> List[ToolRun]:
    hayabusa_report_path = report_folder / f"{folder_name}_hayabusa_output.csv"
    apt_hunter_prefix = report_folder / f"{folder_name}_apt_hunter_output"
    apt_hunter_report_path = Path(str(apt_hunter_prefix) + "_Report.xlsx")
    apt_hunter_additional = [
        Path(str(apt_hunter_prefix) + suffix)
        for suffix in [
            "_TimeSketch.csv",
            "_Logon_Events.csv",
            "_hunting.csv",
        ]
    ]
    chainsaw_report_path = report_folder / f"{folder_name}_chainsaw_output"

    hayabusa_log_path = log_folder / f"{folder_name}_hayabusa.log"
    apt_hunter_log_path = log_folder / f"{folder_name}_apt_hunter.log"
    chainsaw_log_path = log_folder / f"{folder_name}_chainsaw.log"

    hayabusa_cmd = (
        f"{shlex.quote(str(tool_paths['hayabusa']))} csv-timeline --ISO-8601 -t 20 --UTC -q --no-wizard "
        f"-d {shlex.quote(str(target_dir))} -o {shlex.quote(str(hayabusa_report_path))} > {shlex.quote(str(hayabusa_log_path))} 2>&1"
    )

    apt_hunter_cmd = (
        f"{shlex.quote(str(tool_paths['apt_hunter_python']))} {shlex.quote(str(tool_paths['apt_hunter']))} "
        f"-p {shlex.quote(str(target_dir))} -cores 20 -tz UTC -allreport "
        f"-o {shlex.quote(str(apt_hunter_prefix))} > {shlex.quote(str(apt_hunter_log_path))} 2>&1"
    )

    chainsaw_cmd = (
        f"{shlex.quote(str(tool_paths['chainsaw']))} hunt {shlex.quote(str(target_dir))} "
        f"-s {shlex.quote(str(tool_paths['sigma']))} --mapping {shlex.quote(str(tool_paths['mapping']))} "
        f"-r {shlex.quote(str(tool_paths['chainsaw_rules']))} --timezone UTC --full --csv "
        f"-o {shlex.quote(str(chainsaw_report_path))} > {shlex.quote(str(chainsaw_log_path))} 2>&1"
    )

    tool_runs = [
        ToolRun("hayabusa", hayabusa_cmd, hayabusa_log_path, hayabusa_report_path),
        ToolRun(
            "apt-hunter",
            apt_hunter_cmd,
            apt_hunter_log_path,
            apt_hunter_report_path,
            extra_reports=apt_hunter_additional,
        ),
        ToolRun("chainsaw", chainsaw_cmd, chainsaw_log_path, chainsaw_report_path),
    ]

    for tool in tool_runs:
        logging.debug("Prepared %s command: %s", tool.name, tool.command)

    return tool_runs


def launch_tools(tools: List[ToolRun]) -> None:
    for tool in tools:
        logging.debug("Launching %s", tool.name)
        tool.process = subprocess.Popen(tool.command, shell=True)


def monitor_tools(tools: List[ToolRun], progress_interval: float = 5.0) -> None:
    running = {tool.name for tool in tools}
    finished_order = []
    start_time = time.time()
    last_progress_print = 0.0

    previous_snapshot: Dict[str, Tuple[str, Optional[str]]] = {}

    def progress_state(tool: ToolRun, running_flag: bool) -> Tuple[str, Optional[str]]:
        status = "finished" if not running_flag else "running"
        summaries = [s for s in get_artifact_summaries(tool) if s.get("exists")]
        size = None
        if summaries:
            total_bytes = sum(s.get("size_bytes", 0) for s in summaries)
            size = human_size(total_bytes) if total_bytes else None
        return status, size

    while any(tool.process and tool.process.poll() is None for tool in tools):
        current_time = time.time()
        elapsed_time = current_time - start_time

        for tool in tools:
            if tool.process and tool.process.poll() is not None and tool.name in running:
                running.remove(tool.name)
                finished_order.append(tool.name)
                tool.return_code = tool.process.poll()
                print(
                    f"Finished tools: {', '.join(finished_order)} ({elapsed_time:.2f} seconds)"
                )
                logging.debug("%s exited with code %s", tool.name, tool.return_code)

        if current_time - last_progress_print >= progress_interval:
            snapshot: Dict[str, Tuple[str, Optional[str]]] = {}
            for tool in tools:
                snapshot[tool.name] = progress_state(tool, tool.name in running)

            delta_messages = []
            for name, state in snapshot.items():
                if previous_snapshot.get(name) == state:
                    continue
                status, size = state
                if size:
                    message = f"{name} {status} ({size})"
                else:
                    message = f"{name} {status}"

                if status == "finished":
                    message = colorize(message, COLOR_GREEN)
                elif status == "running":
                    message = colorize(message, COLOR_YELLOW)
                delta_messages.append(message)

            if delta_messages:
                timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
                print(colorize(f"[{timestamp}] { ' | '.join(delta_messages) }", COLOR_BLUE))
                previous_snapshot = snapshot
            last_progress_print = current_time

        time.sleep(0.5)

    for tool in tools:
        if tool.process:
            tool.return_code = tool.process.poll()


def rerun_failed_tools(tools: List[ToolRun], enable_retry: bool) -> List[ToolRun]:
    failed = [tool for tool in tools if tool.return_code not in (0, None)]
    if not failed or not enable_retry:
        return failed

    remaining_failures = []
    for tool in failed:
        print(f"Retrying {tool.name}...")
        logging.debug("Retrying command for %s: %s", tool.name, tool.command)
        if tool.log_path.exists():
            try:
                tool.log_path.unlink()
            except OSError:
                pass
        result = subprocess.run(tool.command, shell=True)
        tool.return_code = result.returncode
        if result.returncode != 0:
            remaining_failures.append(tool)
            print(f"{tool.name} failed again with exit code {result.returncode}.")
            logging.debug("%s retry failed with code %s", tool.name, result.returncode)
        else:
            print(f"{tool.name} completed successfully on retry.")
            logging.debug("%s retry succeeded", tool.name)
    return remaining_failures


def summarize_csv(path: Path) -> str:
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            row_count = sum(1 for _ in reader)
        if row_count == 0:
            return "empty CSV"
        return f"{row_count - 1} records"
    except OSError:
        return "unreadable"


def summarize_directory(path: Path) -> str:
    if not path.exists():
        return "missing"
    if path.is_file():
        return human_size(path.stat().st_size)
    files = [f for f in path.rglob("*") if f.is_file()]
    total_size = sum(f.stat().st_size for f in files)
    return f"{len(files)} files ({human_size(total_size)})"


def summarize_tools(tools: List[ToolRun]) -> None:
    print("\nSummary")
    for tool in tools:
        status = "success" if tool.return_code == 0 else f"failed ({tool.return_code})"
        summaries = [s for s in get_artifact_summaries(tool) if s.get("exists")]
        if summaries:
            total_size = sum(s.get("size_bytes", 0) for s in summaries)
            detail = f"{len(summaries)} artifacts ({human_size(total_size)})"
        else:
            detail = "missing"
        log_detail = (
            human_size(tool.log_path.stat().st_size) if tool.log_path.exists() else "missing"
        )
        color = COLOR_GREEN if tool.return_code == 0 else COLOR_RED
        line = f"- {tool.name}: {status}; report {detail}; log {log_detail}"
        print(colorize(line, color))
        logging.debug(
            "%s summary -> status: %s, report detail: %s, log detail: %s",
            tool.name,
            status,
            detail,
            log_detail,
        )


def create_archive(target_dir: Path, folder_name: str, report_folder: Path, log_folder: Path) -> Path:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    archive_path = target_dir / f"{folder_name}_eventlog_bundle_{timestamp}.zip"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for directory in [report_folder, log_folder]:
            if directory.exists():
                for file_path in directory.rglob("*"):
                    if file_path.is_file():
                        arcname = file_path.relative_to(target_dir)
                        zipf.write(file_path, arcname)
    print(f"Packaged artifacts at {archive_path}")
    return archive_path


def ensure_directories(report_folder: Path, log_folder: Path) -> None:
    os.makedirs(report_folder, exist_ok=True)
    os.makedirs(log_folder, exist_ok=True)


def tool_outputs_exist(tools: List[ToolRun]) -> bool:
    for tool in tools:
        for path in [tool.report_path, *tool.extra_reports]:
            if path.exists():
                return True
    return False


def handle_existing_outputs(
    tools: List[ToolRun],
    folder_name: str,
    target_dir: Path,
    report_folder: Path,
    log_folder: Path,
    llm_settings: Dict[str, object],
) -> str:
    menu = (
        "Existing outputs detected for case '{name}'. Choose an action:\n"
        "  [R] Re-run full processing (default)\n"
        "  [L] Re-run LLM summary only\n"
        "  [A] Archive existing outputs\n"
        "  [S] Skip and exit\n"
    ).format(name=folder_name)

    while True:
        print(menu)
        choice = input("Select option [R/L/A/S]: ").strip().lower()
        if choice in {"", "r"}:
            return "rerun"
        if choice == "l":
            for tool in tools:
                tool.return_code = 0
            summary = run_llm_summary(llm_settings, tools, folder_name, report_folder)
            if summary and summary.get("content"):
                print("\n===== LLM Summary =====")
                print(summary["content"])
                print("===== End of LLM Summary =====\n")
            return "exit"
        if choice == "a":
            create_archive(target_dir, folder_name, report_folder, log_folder)
            continue
        if choice == "s":
            return "exit"
        print("Invalid choice. Please select R, L, A, or S.")

def build_llm_context(tools: List[ToolRun], folder_name: str) -> str:
    context: Dict[str, object] = {
        "case": folder_name,
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "tools": [],
    }

    for tool in tools:
        status = "success" if tool.return_code == 0 else f"failed ({tool.return_code})"
        artifacts = get_artifact_summaries(tool)
        context["tools"].append(
            {
                "name": tool.name,
                "status": status,
                "artifacts": artifacts,
                "log": str(tool.log_path),
            }
        )

    return json.dumps(context, indent=2)


def run_llm_summary(
    llm_settings: Dict[str, object],
    tools: List[ToolRun],
    folder_name: str,
    report_folder: Path,
) -> Optional[Dict[str, str]]:
    if requests is None:
        print("LLM summary requested but the 'requests' package is not installed. Run 'pip install requests'.")
        return None

    endpoint = str(llm_settings.get("endpoint", "")).strip()
    if not endpoint:
        print("LLM endpoint is not configured. Update the configuration and retry.")
        return None

    context = build_llm_context(tools, folder_name)
    if not context.strip():
        print("No report artifacts available for LLM summarization.")
        return None

    if not llm_settings.get("enabled", False):
        logging.warning(
            "LLM integration is disabled in configuration, but --llm-summary was requested. Proceeding with stored endpoint."
        )

    system_prompt = llm_settings.get("system_prompt", DEFAULT_LLM_SETTINGS["system_prompt"])
    model = llm_settings.get("model", DEFAULT_LLM_SETTINGS["model"])
    temperature = llm_settings.get("temperature", 0.2)
    max_tokens = llm_settings.get("max_tokens", -1)

    user_prompt = (
        "The following JSON describes outputs from Windows event log analysis tools."
        " Examine the artifacts to surface notable findings, correlations, and any gaps.\n"
        f"Case identifier: {folder_name}.\n"
        "JSON data:\n"
        f"{context}\n\n"
        "Instructions:\n"
        "1. Produce a markdown table with columns: Tool, Artifact, Key Finding, Evidence Snippet, Confidence.\n"
        "   - Use concise entries (<=120 characters per cell).\n"
        "   - When an artifact comes from multiple files (e.g., directory contents), mention each file in Evidence.\n"
        "   - Highlight if an artifact is unreadable or a tool failed.\n"
        "2. Provide a narrative summary (<= 180 words) that: \n"
        "   - Synthesizes correlations between tools.\n"
        "   - Calls out investigative next steps.\n"
        "   - Notes any data quality issues.\n"
        "3. Highlight any notable workbook sheets (e.g., APT-Hunter TimeSketch) with why they matter.\n"
        "4. Add a 'Reasoning Notes' section that explains, in bullet points, how the evidence informed each major finding (reference table rows explicitly).\n"
        "5. Suggest two follow-up questions the analyst should ask."
    )

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "stream": False,
    }
    if isinstance(max_tokens, int) and max_tokens >= 0:
        payload["max_tokens"] = max_tokens

    headers = {"Content-Type": "application/json"}

    try:
        timeout_seconds = llm_settings.get("timeout_seconds", 180)
        response = requests.post(endpoint, headers=headers, json=payload, timeout=timeout_seconds)
        response.raise_for_status()
    except requests.RequestException as exc:  # pragma: no cover - network interaction
        logging.error("LLM summary request failed: %s", exc)
        print(f"LLM summary request failed: {exc}")
        return None

    try:
        data = response.json()
    except json.JSONDecodeError as exc:  # pragma: no cover - network interaction
        logging.error("Failed to decode LLM response JSON: %s", exc)
        print("LLM response was not valid JSON.")
        return None

    content = ""
    choices = data.get("choices")
    if isinstance(choices, list) and choices:
        for choice in choices:
            message = choice.get("message", {})
            role = message.get("role")
            content = message.get("content", "")
            if role == "assistant" and content:
                content = content.strip()
                break

    if not content:
        logging.error("LLM response did not contain summary content: %s", data)
        print("LLM response did not contain summary content.")
        return None

    summary_path = report_folder / f"{folder_name}_llm_summary.txt"
    summary_path.write_text(content, encoding="utf-8")
    print(f"LLM summary saved to {summary_path}")
    return {"path": str(summary_path), "content": content}
def run_update_step(name: str, command: List[str], cwd: Path) -> None:
    logging.info("Updating %s", name)
    logging.debug("Command: %s (cwd=%s)", command, cwd)
    try:
        result = subprocess.run(command, cwd=str(cwd), check=False)
        if result.returncode != 0:
            logging.warning("%s update exited with code %s", name, result.returncode)
        else:
            logging.info("%s update completed", name)
    except FileNotFoundError:
        logging.warning("Command not found while updating %s: %s", name, command[0])


def update_tools(tool_paths: dict) -> None:
    hayabusa_bin = tool_paths["hayabusa"].resolve()
    hayabusa_root = hayabusa_bin.parent

    apt_hunter_path = tool_paths["apt_hunter"].resolve()
    apt_hunter_root = apt_hunter_path.parent

    chainsaw_bin = tool_paths["chainsaw"].resolve()
    chainsaw_root = chainsaw_bin.parents[2] if len(chainsaw_bin.parents) >= 3 else chainsaw_bin.parent

    sigma_root = tool_paths["sigma"].resolve()

    updates = [
        ("Hayabusa", ["git", "pull", "--ff-only"], hayabusa_root),
        ("APT-Hunter", ["git", "pull", "--ff-only"], apt_hunter_root),
        ("Chainsaw", ["git", "pull", "--ff-only"], chainsaw_root),
        ("Sigma", ["git", "pull", "--ff-only"], sigma_root),
    ]

    requirements = apt_hunter_root / "requirements.txt"
    apt_hunter_python = tool_paths["apt_hunter_python"].resolve()
    if requirements.exists():
        if ".venv" not in str(apt_hunter_python):
            logging.warning(
                "Skipping APT-Hunter requirements update because interpreter %s does not look like a virtualenv."
                " Create a venv and point the configuration at it to enable automatic dependency installs.",
                apt_hunter_python,
            )
        else:
            updates.append(
                (
                    "APT-Hunter requirements",
                    [
                        str(apt_hunter_python),
                        "-m",
                        "pip",
                        "install",
                        "-r",
                        str(requirements),
                    ],
                    apt_hunter_root,
                )
            )

    cargo_path = hayabusa_root / "Cargo.toml"
    if cargo_path.exists():
        updates.append(
            (
                "Hayabusa build",
                ["cargo", "build", "--release"],
                hayabusa_root,
            )
        )

    for name, command, cwd in updates:
        if not cwd.exists():
            logging.warning("Skipping %s update; directory missing: %s", name, cwd)
            continue
        run_update_step(name, command, cwd)


def setup_logging(debug: bool, debug_log: Optional[Path]) -> None:
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter(
        "[%(asctime)s] %(levelname)s: %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )
    logger.handlers.clear()

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.DEBUG if debug else logging.WARNING)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    if debug_log:
        debug_log = debug_log.expanduser()
        debug_log.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(debug_log, mode="w", encoding="utf-8")
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    logging.debug("Logging initialized. debug=%s, debug_log=%s", debug, debug_log)


def main() -> None:
    args = parse_args()
    setup_logging(args.debug, args.debug_log)
    target_dir = resolve_target_directory(args)
    folder_name = target_dir.name
    config = load_configuration()
    tool_paths = get_tool_paths(config)
    llm_settings = load_llm_settings(config)
    logging.debug("Using tool paths: %s", tool_paths)

    if args.update_tools:
        update_tools(tool_paths)

    report_folder = target_dir / "eventlog_operations_output"
    log_folder = report_folder / "eventlog_operations_log"
    ensure_directories(report_folder, log_folder)
    logging.debug("Report folder: %s, Log folder: %s", report_folder, log_folder)

    tools = prepare_tools(target_dir, folder_name, report_folder, log_folder, tool_paths)

    if tool_outputs_exist(tools):
        action = handle_existing_outputs(
            tools,
            folder_name,
            target_dir,
            report_folder,
            log_folder,
            llm_settings,
        )
        if action == "exit":
            return

    launch_tools(tools)
    monitor_tools(tools)

    failed = rerun_failed_tools(tools, args.retry_failed)
    summarize_tools(tools)

    if failed:
        print("\nFailures detected:")
        for tool in failed:
            print(f"- {tool.name} (exit code {tool.return_code}). See {tool.log_path}")
    else:
        print("\nAll tools completed successfully.")

    llm_summary = None
    if args.llm_summary:
        llm_summary = run_llm_summary(llm_settings, tools, folder_name, report_folder)

    if llm_summary and llm_summary.get("content"):
        print(colorize("\n===== LLM Summary =====", COLOR_BLUE))
        content = llm_summary["content"]
        lines = content.splitlines()
        for line in lines:
            stripped = line.lstrip()
            if stripped.startswith("|") and stripped.endswith("|"):
                parts = line.split("|")
                if len(parts) >= 3:
                    conf_index = len(parts) - 2
                    confidence_cell = parts[conf_index]
                    conf_value = confidence_cell.strip()
                    colored_conf = colorize_confidence(conf_value)
                    parts[conf_index] = confidence_cell.replace(conf_value, colored_conf)
                table_line = "|".join(parts)
                if "Tool" in stripped or stripped.replace("-", "").strip() == "":
                    print(colorize(table_line, COLOR_BLUE))
                else:
                    print(table_line)
            elif stripped.startswith("#"):
                print(colorize(line, COLOR_BLUE))
            elif stripped.startswith("-") or stripped.startswith("*"):
                print(colorize(line, COLOR_GREEN))
            else:
                print(colorize(line, COLOR_GREEN))
        print(colorize("===== End of LLM Summary =====\n", COLOR_BLUE))

    if args.archive:
        create_archive(target_dir, folder_name, report_folder, log_folder)

    input("Press Enter to exit...")


if __name__ == "__main__":
    main()
