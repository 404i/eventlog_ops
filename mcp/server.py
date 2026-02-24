"""
eventlog-ops MCP Server
=======================
Exposes the outputs from eventlog_ops (Hayabusa, APT-Hunter, Chainsaw) as
MCP tools so Claude can act as a DFIR analyst against a processed case.

Start:
    python -m mcp.server  (see requirements.txt)
    OR via claude_desktop_config.json  (see claude_config_example.json)

Read-only query tools (1-10) are stateless – every call reads from the
filesystem so results always reflect the current on-disk state of the case
output directory.

Execution tools (11-13):
  run_analysis   – launch eventlog_operations_v4.py against an EVTX directory
  get_run_status – poll a running job's progress and output tail
  cancel_run     – terminate a running job
"""

from __future__ import annotations

import csv
import glob
import json
import os
import re
import signal
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

from mcp.server.fastmcp import FastMCP

mcp = FastMCP(
    "eventlog-ops-dfir",
    instructions=(
        "You are a DFIR analyst assistant with direct access to the local Mac "
        "filesystem via this MCP server. "
        "IMPORTANT: Never use bash or shell tools to run eventlog_ops. "
        "Always use the run_analysis MCP tool to execute the analysis — it runs "
        "directly on the local Mac, can reach any local path, and returns a job_id "
        "you can poll with get_run_status. "
        "After analysis completes, use query_hayabusa, query_chainsaw, "
        "query_apthunter, get_merged_timeline and search_all to investigate "
        "the results. Always cite the specific artifact (file, row, rule name) "
        "behind each finding."
    ),
)

# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}

OUTPUT_SUBDIR = "eventlog_operations_output"
LOG_SUBDIR = "eventlog_operations_log"

# Hayabusa CSV column names (csv-timeline --ISO-8601 output)
HAY_COLS = [
    "Datetime", "TimestampDesc", "RuleTitle", "Level",
    "Computer", "Channel", "EventID", "MitreTactics", "MitreTags",
    "OtherTags", "RecordID", "Details", "ExtraFieldInfo", "RuleFile", "EvtxFile",
]


def _output_dir(case_dir: str) -> Path:
    """Return the eventlog_operations_output directory for a case."""
    p = Path(case_dir).expanduser().resolve()
    out = p / OUTPUT_SUBDIR
    if out.is_dir():
        return out
    # Caller may have passed the output dir directly
    if (p / "eventlog_operations_log").is_dir():
        return p
    raise FileNotFoundError(
        f"No '{OUTPUT_SUBDIR}' directory found under '{case_dir}'. "
        "Pass the parent of that directory (the case/EVTX root)."
    )


def _case_name(case_dir: str) -> str:
    return Path(case_dir).expanduser().resolve().name


def _hayabusa_csv(case_dir: str) -> Optional[Path]:
    out = _output_dir(case_dir)
    # Try canonical name first, then glob for any prefix (handles Docker mount naming)
    name = out.parent.name
    p = out / f"{name}_hayabusa_output.csv"
    if p.exists():
        return p
    matches = sorted(out.glob("*_hayabusa_output.csv"))
    return matches[0] if matches else None


def _apthunter_xlsx(case_dir: str) -> Optional[Path]:
    out = _output_dir(case_dir)
    name = out.parent.name
    p = out / f"{name}_apt_hunter_output_Report.xlsx"
    if p.exists():
        return p
    matches = sorted(out.glob("*_apt_hunter_output_Report.xlsx"))
    return matches[0] if matches else None


def _apthunter_csvs(case_dir: str) -> Dict[str, Path]:
    """Return dict of suffix → Path for all APT-Hunter CSV outputs."""
    out = _output_dir(case_dir)
    name = out.parent.name
    result: Dict[str, Path] = {}
    for suffix in ["TimeSketch", "Logon_Events", "hunting"]:
        p = out / f"{name}_apt_hunter_output_{suffix}.csv"
        if p.exists():
            result[suffix] = p
        else:
            matches = sorted(out.glob(f"*_apt_hunter_output_{suffix}.csv"))
            if matches:
                result[suffix] = matches[0]
    return result


def _chainsaw_dir(case_dir: str) -> Optional[Path]:
    out = _output_dir(case_dir)
    name = out.parent.name
    p = out / f"{name}_chainsaw_output"
    if p.is_dir():
        return p
    matches = sorted(out.glob("*_chainsaw_output"))
    return matches[0] if matches else None


def _llm_summary(case_dir: str) -> Optional[Path]:
    out = _output_dir(case_dir)
    name = out.parent.name
    p = out / f"{name}_llm_summary.txt"
    if p.exists():
        return p
    matches = sorted(out.glob("*_llm_summary.txt"))
    return matches[0] if matches else None


def _log_path(case_dir: str, tool: str) -> Optional[Path]:
    out = _output_dir(case_dir)
    log_dir = out / LOG_SUBDIR
    name = out.parent.name
    p = log_dir / f"{name}_{tool}.log"
    if p.exists():
        return p
    matches = sorted(log_dir.glob(f"*_{tool}.log"))
    return matches[0] if matches else None


def _read_csv_rows(path: Path, limit: int) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= limit:
                break
    return rows


def _match_row(row: Dict[str, str], filters: Dict[str, Optional[str]]) -> bool:
    """Return True if all non-None filters match (case-insensitive substring)."""
    for col, val in filters.items():
        if val is None:
            continue
        cell = str(row.get(col, "")).lower()
        if val.lower() not in cell:
            return False
    return True


def _time_in_range(dt_str: str, start: Optional[str], end: Optional[str]) -> bool:
    if not start and not end:
        return True
    # Accept ISO-8601 prefix comparison (string sort works for ISO dates)
    s = dt_str[:19]  # "2024-01-15T12:34:56"
    if start and s < start[:19]:
        return False
    if end and s > end[:19]:
        return False
    return True


def _severity_lte(level: str, max_level: Optional[str]) -> bool:
    if max_level is None:
        return True
    return SEVERITY_ORDER.get(level.lower(), 99) <= SEVERITY_ORDER.get(max_level.lower(), 99)


def _severity_gte(level: str, min_level: Optional[str]) -> bool:
    if min_level is None:
        return True
    return SEVERITY_ORDER.get(level.lower(), 99) >= SEVERITY_ORDER.get(min_level.lower(), 99)


def _truncate(rows: List[Any], limit: int) -> Dict[str, Any]:
    total = len(rows)
    return {
        "total_matched": total,
        "returned": min(total, limit),
        "truncated": total > limit,
        "rows": rows[:limit],
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 1 – list_cases
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def list_cases(base_dir: str) -> str:
    """
    Scan a directory for processed eventlog_ops cases.

    Args:
        base_dir: Root directory to search (e.g. ~/cases or /data).

    Returns:
        JSON list of case directories that contain eventlog_operations_output.
    """
    base = Path(base_dir).expanduser().resolve()
    if not base.is_dir():
        return json.dumps({"error": f"'{base_dir}' is not a directory."})

    cases = []
    for child in sorted(base.iterdir()):
        out = child / OUTPUT_SUBDIR
        if out.is_dir():
            # Gather quick stats
            artifacts = list(out.glob("*"))
            cases.append({
                "case_dir": str(child),
                "case_name": child.name,
                "artifact_count": len(artifacts),
                "has_hayabusa": any("hayabusa" in f.name for f in artifacts),
                "has_apt_hunter": any("apt_hunter" in f.name for f in artifacts),
                "has_chainsaw": any("chainsaw" in f.name for f in artifacts),
                "has_llm_summary": any("llm_summary" in f.name for f in artifacts),
            })

    return json.dumps({"base_dir": str(base), "cases": cases}, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 2 – get_case_summary
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_case_summary(case_dir: str) -> str:
    """
    Return a structured overview of all tool outputs for a case:
    file sizes, row counts, Hayabusa severity breakdown, and APT-Hunter sheet list.

    Args:
        case_dir: Path to the case root (parent of eventlog_operations_output/).
    """
    try:
        out = _output_dir(case_dir)
    except FileNotFoundError as e:
        return json.dumps({"error": str(e)})

    name = out.parent.name
    summary: Dict[str, Any] = {"case": name, "output_dir": str(out), "tools": {}}

    # ── Hayabusa ──
    hay = _hayabusa_csv(case_dir)
    if hay:
        severity_counts: Dict[str, int] = {}
        computers: set = set()
        row_count = 0
        with hay.open("r", encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                row_count += 1
                lvl = row.get("Level", "unknown").lower()
                severity_counts[lvl] = severity_counts.get(lvl, 0) + 1
                comp = row.get("Computer", "")
                if comp:
                    computers.add(comp)
        summary["tools"]["hayabusa"] = {
            "file": hay.name,
            "size_bytes": hay.stat().st_size,
            "total_events": row_count,
            "severity_breakdown": dict(sorted(severity_counts.items(), key=lambda x: SEVERITY_ORDER.get(x[0], 99))),
            "unique_computers": sorted(computers),
        }
    else:
        summary["tools"]["hayabusa"] = {"status": "not found"}

    # ── APT-Hunter ──
    apt_xlsx = _apthunter_xlsx(case_dir)
    apt_entry: Dict[str, Any] = {}
    if apt_xlsx:
        apt_entry["report_xlsx"] = {"file": apt_xlsx.name, "size_bytes": apt_xlsx.stat().st_size}
        if load_workbook:
            try:
                wb = load_workbook(apt_xlsx, read_only=True, data_only=True)
                apt_entry["report_xlsx"]["sheets"] = {
                    sheet: wb[sheet].max_row for sheet in wb.sheetnames
                }
                wb.close()
            except Exception as e:
                apt_entry["report_xlsx"]["error"] = str(e)
    else:
        apt_entry["report_xlsx"] = {"status": "not found"}

    apt_csvs = _apthunter_csvs(case_dir)
    apt_entry["csv_outputs"] = {
        suffix: {"file": p.name, "size_bytes": p.stat().st_size}
        for suffix, p in apt_csvs.items()
    }
    summary["tools"]["apt_hunter"] = apt_entry

    # ── Chainsaw ──
    cs_dir = _chainsaw_dir(case_dir)
    if cs_dir:
        cs_files = list(cs_dir.rglob("*.csv"))
        cs_entry: Dict[str, Any] = {
            "output_dir": cs_dir.name,
            "csv_files": len(cs_files),
            "groups": {},
        }
        for f in cs_files:
            try:
                with f.open("r", encoding="utf-8", errors="replace") as fh:
                    count = sum(1 for _ in fh) - 1  # subtract header
                cs_entry["groups"][f.stem] = {"file": f.name, "events": max(count, 0)}
            except OSError:
                pass
        summary["tools"]["chainsaw"] = cs_entry
    else:
        summary["tools"]["chainsaw"] = {"status": "not found"}

    # ── LLM summary ──
    llm = _llm_summary(case_dir)
    summary["tools"]["llm_summary"] = {
        "available": llm is not None,
        "file": llm.name if llm else None,
        "size_bytes": llm.stat().st_size if llm else None,
    }

    return json.dumps(summary, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 3 – query_hayabusa
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def query_hayabusa(
    case_dir: str,
    min_severity: Optional[str] = None,
    rule_contains: Optional[str] = None,
    computer: Optional[str] = None,
    channel: Optional[str] = None,
    event_id: Optional[str] = None,
    mitre_tactic: Optional[str] = None,
    details_contains: Optional[str] = None,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    limit: int = 100,
) -> str:
    """
    Query and filter the Hayabusa CSV-timeline output.

    Args:
        case_dir:         Path to the case root.
        min_severity:     Minimum severity to include: critical / high / medium / low / informational.
        rule_contains:    Substring to match against RuleTitle.
        computer:         Substring to match against Computer field.
        channel:          Substring to match against Channel (e.g. 'Security', 'System').
        event_id:         Exact or partial EventID to match.
        mitre_tactic:     Substring to match in MitreTactics (e.g. 'Lateral').
        details_contains: Substring to search in the Details field.
        start_time:       ISO-8601 start datetime (e.g. '2024-01-15T10:00:00').
        end_time:         ISO-8601 end datetime (e.g. '2024-01-15T18:00:00').
        limit:            Maximum rows to return (default 100, max 1000).
    """
    limit = min(max(1, limit), 1000)
    hay = _hayabusa_csv(case_dir)
    if not hay:
        return json.dumps({"error": "Hayabusa output CSV not found for this case."})

    filters = {
        "RuleTitle": rule_contains,
        "Computer": computer,
        "Channel": channel,
        "EventID": event_id,
        "MitreTactics": mitre_tactic,
        "Details": details_contains,
    }

    matched: List[Dict[str, str]] = []
    with hay.open("r", encoding="utf-8", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if not _match_row(row, filters):
                continue
            if not _time_in_range(row.get("Datetime", ""), start_time, end_time):
                continue
            if not _severity_gte(row.get("Level", ""), min_severity):
                continue
            matched.append(row)

    result = _truncate(matched, limit)
    result["source_file"] = hay.name
    result["filters_applied"] = {k: v for k, v in filters.items() if v is not None}
    return json.dumps(result, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 4 – query_chainsaw
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def query_chainsaw(
    case_dir: str,
    group: Optional[str] = None,
    rule_contains: Optional[str] = None,
    computer: Optional[str] = None,
    event_id: Optional[str] = None,
    level: Optional[str] = None,
    details_contains: Optional[str] = None,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    limit: int = 100,
) -> str:
    """
    Query Chainsaw CSV outputs. Each Sigma detection group is a separate CSV file.

    Args:
        case_dir:         Path to the case root.
        group:            CSV file stem / detection group name (substring match).
                          Pass None to search across all groups.
        rule_contains:    Substring to match against the rule name column.
        computer:         Substring to match against the computer/system column.
        event_id:         Partial EventID match.
        level:            Severity level (substring: critical/high/medium/low).
        details_contains: Substring to search across all columns (full-row search).
        start_time:       ISO-8601 start datetime.
        end_time:         ISO-8601 end datetime.
        limit:            Maximum rows to return (default 100, max 1000).
    """
    limit = min(max(1, limit), 1000)
    cs_dir = _chainsaw_dir(case_dir)
    if not cs_dir:
        return json.dumps({"error": "Chainsaw output directory not found for this case."})

    csv_files = list(cs_dir.rglob("*.csv"))
    if group:
        csv_files = [f for f in csv_files if group.lower() in f.stem.lower()]

    if not csv_files:
        return json.dumps({
            "error": f"No Chainsaw CSV files found" + (f" matching group '{group}'." if group else "."),
            "available_groups": [f.stem for f in sorted(cs_dir.rglob("*.csv"))],
        })

    matched: List[Dict[str, Any]] = []
    for csv_file in sorted(csv_files):
        try:
            with csv_file.open("r", encoding="utf-8", errors="replace") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    row_flat = dict(row)
                    row_flat["_group"] = csv_file.stem

                    # Time filter – try common timestamp columns
                    ts = (
                        row_flat.get("system_time")
                        or row_flat.get("timestamp")
                        or row_flat.get("Timestamp")
                        or row_flat.get("SystemTime")
                        or ""
                    )
                    if not _time_in_range(ts, start_time, end_time):
                        continue

                    # Specific field filters
                    if computer:
                        comp_val = (
                            row_flat.get("computer")
                            or row_flat.get("Computer")
                            or row_flat.get("system.computer")
                            or ""
                        )
                        if computer.lower() not in comp_val.lower():
                            continue

                    if event_id:
                        eid_val = (
                            row_flat.get("event_id")
                            or row_flat.get("EventID")
                            or row_flat.get("event.event.code")
                            or ""
                        )
                        if event_id not in str(eid_val):
                            continue

                    if level:
                        lvl_val = (
                            row_flat.get("level")
                            or row_flat.get("Level")
                            or ""
                        )
                        if level.lower() not in lvl_val.lower():
                            continue

                    if rule_contains:
                        name_val = (
                            row_flat.get("name")
                            or row_flat.get("Name")
                            or row_flat.get("rule_title")
                            or ""
                        )
                        if rule_contains.lower() not in name_val.lower():
                            continue

                    if details_contains:
                        full_row_text = " ".join(str(v) for v in row_flat.values()).lower()
                        if details_contains.lower() not in full_row_text:
                            continue

                    matched.append(row_flat)
        except OSError:
            pass

    result = _truncate(matched, limit)
    result["groups_searched"] = [f.stem for f in sorted(csv_files)]
    return json.dumps(result, indent=2, default=str)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 5 – query_apthunter
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def query_apthunter(
    case_dir: str,
    source: str = "Report",
    sheet: Optional[str] = None,
    keyword: Optional[str] = None,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    limit: int = 100,
) -> str:
    """
    Query APT-Hunter outputs (XLSX report or CSV outputs like TimeSketch, Logon_Events, hunting).

    Args:
        case_dir:   Path to the case root.
        source:     Which output to query:
                    'Report' (XLSX), 'TimeSketch', 'Logon_Events', or 'hunting'.
        sheet:      XLSX sheet name to read (only used when source='Report').
                    Pass None to see the sheet list first.
        keyword:    Substring to match anywhere in a row.
        start_time: ISO-8601 start datetime filter.
        end_time:   ISO-8601 end datetime filter.
        limit:      Maximum rows to return (default 100, max 1000).
    """
    limit = min(max(1, limit), 1000)

    if source == "Report":
        xlsx = _apthunter_xlsx(case_dir)
        if not xlsx:
            return json.dumps({"error": "APT-Hunter XLSX report not found."})
        if load_workbook is None:
            return json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"})
        try:
            wb = load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:
            return json.dumps({"error": f"Failed to open XLSX: {e}"})

        if sheet is None:
            sheets_info = {}
            for sname in wb.sheetnames:
                ws = wb[sname]
                sheets_info[sname] = ws.max_row
            wb.close()
            return json.dumps({
                "source": xlsx.name,
                "available_sheets": sheets_info,
                "hint": "Pass sheet='<name>' to read a specific sheet.",
            }, indent=2)

        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({
                "error": f"Sheet '{sheet}' not found.",
                "available_sheets": wb.sheetnames,
            })

        ws = wb[sheet]
        header: List[str] = []
        rows: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                continue
            row_dict = {header[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(header)}

            # Time filter
            time_filtered = False
            for time_col in ("Date", "Time", "Timestamp", "DateTime", "date", "time"):
                if time_col in row_dict:
                    if not _time_in_range(row_dict[time_col], start_time, end_time):
                        time_filtered = True
                    break  # only check the first matching time column

            if time_filtered:
                continue

            if keyword:
                row_text = " ".join(str(v) for v in row_dict.values()).lower()
                if keyword.lower() not in row_text:
                    continue

            rows.append(row_dict)
        wb.close()

        result = _truncate(rows, limit)
        result["source"] = xlsx.name
        result["sheet"] = sheet
        return json.dumps(result, indent=2, default=str)

    else:
        # CSV-based outputs: TimeSketch, Logon_Events, hunting
        apt_csvs = _apthunter_csvs(case_dir)
        if source not in apt_csvs:
            return json.dumps({
                "error": f"APT-Hunter '{source}' CSV not found.",
                "available_sources": ["Report"] + list(apt_csvs.keys()),
            })

        csv_path = apt_csvs[source]
        rows = []
        with csv_path.open("r", encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                if keyword:
                    row_text = " ".join(str(v) for v in row.values()).lower()
                    if keyword.lower() not in row_text:
                        continue
                rows.append(dict(row))

        result = _truncate(rows, limit)
        result["source"] = csv_path.name
        return json.dumps(result, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 6 – get_merged_timeline
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_merged_timeline(
    case_dir: str,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    min_severity: Optional[str] = None,
    tools: str = "all",
    limit: int = 200,
) -> str:
    """
    Merge detections from Hayabusa, Chainsaw, and the APT-Hunter TimeSketch CSV
    into a single chronological timeline.

    Args:
        case_dir:     Path to the case root.
        start_time:   ISO-8601 start datetime filter.
        end_time:     ISO-8601 end datetime filter.
        min_severity: Minimum severity (critical/high/medium/low/informational).
        tools:        Comma-separated list of tools to include: 'hayabusa', 'chainsaw', 'apthunter'.
                      Default 'all' includes all three.
        limit:        Maximum events to return after sorting (default 200, max 2000).
    """
    limit = min(max(1, limit), 2000)
    include = set(t.strip().lower() for t in tools.split(",")) if tools != "all" else {"hayabusa", "chainsaw", "apthunter"}

    events: List[Dict[str, Any]] = []

    # ── Hayabusa ──
    if "hayabusa" in include:
        hay = _hayabusa_csv(case_dir)
        if hay:
            with hay.open("r", encoding="utf-8", errors="replace") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    ts = row.get("Datetime", "")
                    if not _time_in_range(ts, start_time, end_time):
                        continue
                    if not _severity_gte(row.get("Level", ""), min_severity):
                        continue
                    events.append({
                        "timestamp": ts,
                        "tool": "hayabusa",
                        "level": row.get("Level", ""),
                        "title": row.get("RuleTitle", ""),
                        "computer": row.get("Computer", ""),
                        "event_id": row.get("EventID", ""),
                        "channel": row.get("Channel", ""),
                        "mitre_tactics": row.get("MitreTactics", ""),
                        "details": row.get("Details", "")[:300],
                    })

    # ── Chainsaw ──
    if "chainsaw" in include:
        cs_dir = _chainsaw_dir(case_dir)
        if cs_dir:
            for csv_file in sorted(cs_dir.rglob("*.csv")):
                try:
                    with csv_file.open("r", encoding="utf-8", errors="replace") as fh:
                        reader = csv.DictReader(fh)
                        for row in reader:
                            ts = (
                                row.get("system_time")
                                or row.get("timestamp")
                                or row.get("Timestamp")
                                or ""
                            )
                            if not _time_in_range(ts, start_time, end_time):
                                continue
                            level = row.get("level") or row.get("Level") or ""
                            if min_severity and not _severity_gte(level, min_severity):
                                continue
                            events.append({
                                "timestamp": ts,
                                "tool": "chainsaw",
                                "level": level,
                                "title": row.get("name") or row.get("Name") or csv_file.stem,
                                "computer": row.get("computer") or row.get("Computer") or "",
                                "event_id": row.get("event_id") or row.get("EventID") or "",
                                "channel": row.get("channel") or row.get("Channel") or "",
                                "mitre_tactics": row.get("tags") or row.get("Tags") or "",
                                "details": str(row)[:300],
                                "_group": csv_file.stem,
                            })
                except OSError:
                    pass

    # ── APT-Hunter TimeSketch ──
    if "apthunter" in include:
        apt_csvs = _apthunter_csvs(case_dir)
        ts_csv = apt_csvs.get("TimeSketch")
        if ts_csv:
            with ts_csv.open("r", encoding="utf-8", errors="replace") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    ts = row.get("datetime") or row.get("Datetime") or row.get("Time") or ""
                    if not _time_in_range(ts, start_time, end_time):
                        continue
                    events.append({
                        "timestamp": ts,
                        "tool": "apthunter",
                        "level": row.get("label") or "medium",
                        "title": row.get("message") or row.get("Message") or row.get("description") or "",
                        "computer": row.get("hostname") or row.get("computer") or "",
                        "event_id": row.get("event_id") or "",
                        "channel": row.get("source") or "",
                        "mitre_tactics": "",
                        "details": str(row)[:300],
                    })

    # Sort chronologically
    events.sort(key=lambda e: e.get("timestamp", ""))

    result = _truncate(events, limit)
    result["time_range"] = {"start": start_time, "end": end_time}
    result["tools_included"] = sorted(include)
    return json.dumps(result, indent=2, default=str)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 7 – search_all
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def search_all(
    case_dir: str,
    keyword: str,
    tools: str = "all",
    limit: int = 100,
) -> str:
    """
    Full-text keyword search across ALL tool outputs for a case.
    Useful for hunting a specific username, IP, hash, hostname, or IOC.

    Args:
        case_dir: Path to the case root.
        keyword:  String to search for (case-insensitive).
        tools:    Comma-separated: 'hayabusa', 'chainsaw', 'apthunter'. Default 'all'.
        limit:    Maximum hits to return per tool (default 100, max 500).
    """
    limit = min(max(1, limit), 500)
    include = set(t.strip().lower() for t in tools.split(",")) if tools != "all" else {"hayabusa", "chainsaw", "apthunter"}
    kw = keyword.lower()
    hits: Dict[str, List[Dict]] = {}

    if "hayabusa" in include:
        hay = _hayabusa_csv(case_dir)
        if hay:
            hay_hits: List[Dict] = []
            with hay.open("r", encoding="utf-8", errors="replace") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    if kw in " ".join(str(v) for v in row.values()).lower():
                        hay_hits.append(row)
                        if len(hay_hits) >= limit:
                            break
            hits["hayabusa"] = hay_hits

    if "chainsaw" in include:
        cs_dir = _chainsaw_dir(case_dir)
        cs_hits: List[Dict] = []
        if cs_dir:
            for csv_file in sorted(cs_dir.rglob("*.csv")):
                try:
                    with csv_file.open("r", encoding="utf-8", errors="replace") as fh:
                        reader = csv.DictReader(fh)
                        for row in reader:
                            if kw in " ".join(str(v) for v in row.values()).lower():
                                enriched = dict(row)
                                enriched["_group"] = csv_file.stem
                                cs_hits.append(enriched)
                                if len(cs_hits) >= limit:
                                    break
                except OSError:
                    pass
                if len(cs_hits) >= limit:
                    break
        hits["chainsaw"] = cs_hits

    if "apthunter" in include:
        apt_hits: List[Dict] = []
        apt_csvs = _apthunter_csvs(case_dir)
        for suffix, path in apt_csvs.items():
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    if kw in " ".join(str(v) for v in row.values()).lower():
                        enriched = dict(row)
                        enriched["_source"] = suffix
                        apt_hits.append(enriched)
                        if len(apt_hits) >= limit:
                            break
            if len(apt_hits) >= limit:
                break
        hits["apthunter"] = apt_hits

    totals = {tool: len(rows) for tool, rows in hits.items()}
    return json.dumps({
        "keyword": keyword,
        "totals": totals,
        "results": hits,
    }, indent=2, default=str)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 8 – get_unique_values
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_unique_values(
    case_dir: str,
    tool: str,
    column: str,
    min_count: int = 1,
) -> str:
    """
    Return unique values and their frequency for a specific column in a tool's output.
    Ideal for scoping: 'what computers are in scope?', 'which rule titles fired?', etc.

    Args:
        case_dir:  Path to the case root.
        tool:      'hayabusa', 'chainsaw' (searches all group CSVs), or 'apthunter_timesketch'.
        column:    Column name to aggregate (case-sensitive, use get_case_summary to find column names).
        min_count: Minimum occurrence count to include in results (default 1).
    """
    counts: Dict[str, int] = {}

    if tool == "hayabusa":
        hay = _hayabusa_csv(case_dir)
        if not hay:
            return json.dumps({"error": "Hayabusa CSV not found."})
        with hay.open("r", encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                val = str(row.get(column, "")).strip()
                if val:
                    counts[val] = counts.get(val, 0) + 1

    elif tool == "chainsaw":
        cs_dir = _chainsaw_dir(case_dir)
        if not cs_dir:
            return json.dumps({"error": "Chainsaw output directory not found."})
        for csv_file in sorted(cs_dir.rglob("*.csv")):
            try:
                with csv_file.open("r", encoding="utf-8", errors="replace") as fh:
                    reader = csv.DictReader(fh)
                    for row in reader:
                        val = str(row.get(column, "")).strip()
                        if val:
                            counts[val] = counts.get(val, 0) + 1
            except OSError:
                pass

    elif tool == "apthunter_timesketch":
        apt_csvs = _apthunter_csvs(case_dir)
        ts = apt_csvs.get("TimeSketch")
        if not ts:
            return json.dumps({"error": "APT-Hunter TimeSketch CSV not found."})
        with ts.open("r", encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                val = str(row.get(column, "")).strip()
                if val:
                    counts[val] = counts.get(val, 0) + 1

    else:
        return json.dumps({"error": f"Unknown tool '{tool}'. Choose: hayabusa, chainsaw, apthunter_timesketch."})

    filtered = {v: c for v, c in counts.items() if c >= min_count}
    sorted_counts = dict(sorted(filtered.items(), key=lambda x: x[1], reverse=True))
    return json.dumps({
        "tool": tool,
        "column": column,
        "distinct_values": len(sorted_counts),
        "values": sorted_counts,
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 9 – get_llm_summary
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_llm_summary(case_dir: str) -> str:
    """
    Return the pre-generated LLM summary for a case (if one exists from a
    previous eventlog_ops run with --llm-summary).

    Args:
        case_dir: Path to the case root.
    """
    llm = _llm_summary(case_dir)
    if not llm:
        return json.dumps({
            "available": False,
            "hint": "Run eventlog_operations_v4.py with --llm-summary to generate one.",
        })
    return json.dumps({
        "available": True,
        "file": llm.name,
        "content": llm.read_text(encoding="utf-8", errors="replace"),
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 10 – get_tool_log
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_tool_log(
    case_dir: str,
    tool: str,
    tail: int = 100,
) -> str:
    """
    Return the execution log for a specific tool from the last eventlog_ops run.
    Useful for diagnosing tool failures or understanding what rules were skipped.

    Args:
        case_dir: Path to the case root.
        tool:     Tool name: 'hayabusa', 'apt_hunter', or 'chainsaw'.
        tail:     Number of lines from the end of the log to return (default 100).
    """
    log = _log_path(case_dir, tool)
    if not log:
        return json.dumps({
            "error": f"Log file for '{tool}' not found.",
            "hint": "Valid tool names: hayabusa, apt_hunter, chainsaw",
        })

    lines = log.read_text(encoding="utf-8", errors="replace").splitlines()
    shown = lines[-tail:] if tail < len(lines) else lines
    return json.dumps({
        "tool": tool,
        "log_file": log.name,
        "total_lines": len(lines),
        "showing_last": len(shown),
        "content": "\n".join(shown),
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Job registry  (in-memory; lives as long as the MCP server process)
# ─────────────────────────────────────────────────────────────────────────────

_JOBS: Dict[str, Dict[str, Any]] = {}
_JOBS_LOCK = threading.Lock()
_MAX_FINISHED_JOBS = 50  # max completed/failed/cancelled entries retained

# Likely locations of the main script, checked in order.
_SCRIPT_CANDIDATES = [
    Path(__file__).parent.parent / "eventlog_operations_v4.py",  # dev checkout
    Path("/app/eventlog_operations_v4.py"),                       # Docker image
]


def _find_script(override: Optional[str] = None) -> Optional[Path]:
    if override:
        p = Path(override).expanduser().resolve()
        return p if p.is_file() else None
    for candidate in _SCRIPT_CANDIDATES:
        if candidate.is_file():
            return candidate
    return None


def _stream_output(stream, lines: List[str], prefix: str = "") -> None:
    """Thread target: read lines from *stream* and append to *lines*."""
    try:
        for raw in stream:
            line = raw.rstrip("\n")
            if prefix:
                line = f"[{prefix}] {line}"
            with _JOBS_LOCK:
                lines.append(line)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 11 – run_analysis
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def run_analysis(
    case_dir: str,
    retry_failed: bool = False,
    skip_hayabusa: bool = False,
    skip_chainsaw: bool = False,
    skip_apthunter: bool = False,
    llm_summary: bool = False,
    tool_timeout: Optional[int] = None,
    archive: bool = False,
    script_path: Optional[str] = None,
) -> str:
    """
    Launch eventlog_operations_v4.py against a directory of EVTX files.
    This tool runs DIRECTLY on the local machine via the MCP server — do NOT
    use bash or shell to run it.  The analysis runs in the background; use
    get_run_status(job_id) to poll progress and cancel_run(job_id) to abort.

    Args:
        case_dir:      Path to the directory containing EVTX files.
                       Any local path is accessible (MCP server runs on the Mac).
        retry_failed:  Re-run tools that previously produced no output.
        skip_hayabusa: Skip the Hayabusa step.
        skip_chainsaw: Skip the Chainsaw step.
        skip_apthunter: Skip the APT-Hunter step.
        llm_summary:   Generate an LLM summary after tool runs (requires a
                       configured LLM endpoint in the ini).
        tool_timeout:  Per-tool timeout in seconds (0 = no limit).
        archive:       Create a zip archive of the output when finished.
        script_path:   Override path to eventlog_operations_v4.py. Auto-detected
                       if not supplied.

    Returns:
        JSON with job_id, status, and the resolved case_dir.
    """
    target = Path(case_dir).expanduser().resolve()
    if not target.is_dir():
        return json.dumps({"error": f"'{case_dir}' is not a directory."})

    # ── Decide: Docker (preferred) or native script fallback ─────────────────
    use_docker = script_path is None  # explicit script_path forces native mode
    docker_image = "eventlog-ops:latest"

    if use_docker:
        # Verify the image exists before committing
        check = subprocess.run(
            ["docker", "image", "inspect", docker_image],
            capture_output=True, text=True,
        )
        if check.returncode != 0:
            use_docker = False  # fall through to native

    if use_docker:
        # docker run --rm  -v <case_dir>:/data  eventlog-ops:latest  [flags] /data
        cmd: List[str] = [
            "docker", "run", "--rm",
            "--volume", f"{target}:/data",
            "--env", "EVENTLOG_AUTO_CONFIG=1",
            "--env", "PYTHONUNBUFFERED=1",
        ]
        if llm_summary:
            # Pass LLM env vars through if set on the host
            for var in ("LLM_ENDPOINT", "LLM_MODEL"):
                val = os.environ.get(var, "")
                if val:
                    cmd += ["--env", f"{var}={val}"]
        cmd.append(docker_image)
        # Extra flags for the script (entrypoint handles /data automatically)
        if retry_failed:
            cmd.append("--retry-failed")
        if skip_hayabusa:
            cmd.append("--skip-hayabusa")
        if skip_chainsaw:
            cmd.append("--skip-chainsaw")
        if skip_apthunter:
            cmd.append("--skip-apt-hunter")
        if llm_summary:
            cmd.append("--llm-summary")
        if archive:
            cmd.append("--archive")
        if tool_timeout is not None:
            cmd += ["--tool-timeout", str(tool_timeout)]
        cmd.append("/data")
        env = os.environ.copy()
    else:
        # Native fallback – script must be reachable and tools installed locally
        script = _find_script(script_path)
        if script is None:
            searched = ", ".join(str(c) for c in _SCRIPT_CANDIDATES)
            return json.dumps({
                "error": (
                    f"Docker image '{docker_image}' not found and "
                    "eventlog_operations_v4.py not found for native fallback."
                ),
                "searched": searched,
                "hint": (
                    "Build the image first: "
                    "docker compose -f <repo>/docker/docker-compose.yml build"
                ),
            })
        cmd = [sys.executable, str(script), str(target), "--auto-config"]
        if retry_failed:
            cmd.append("--retry-failed")
        if skip_hayabusa:
            cmd.append("--skip-hayabusa")
        if skip_chainsaw:
            cmd.append("--skip-chainsaw")
        if skip_apthunter:
            cmd.append("--skip-apt-hunter")
        if llm_summary:
            cmd.append("--llm-summary")
        if archive:
            cmd.append("--archive")
        if tool_timeout is not None:
            cmd += ["--tool-timeout", str(tool_timeout)]
        env = os.environ.copy()
        env["EVENTLOG_AUTO_CONFIG"] = "1"
        env["PYTHONUNBUFFERED"] = "1"

    job_id = str(uuid.uuid4())[:8]
    output_lines: List[str] = []

    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
            env=env,
        )
    except Exception as e:
        return json.dumps({"error": f"Failed to start process: {e}"})

    # Drain stdout and stderr in background threads so the pipe never blocks.
    t_out = threading.Thread(
        target=_stream_output, args=(proc.stdout, output_lines, "OUT"), daemon=True
    )
    t_err = threading.Thread(
        target=_stream_output, args=(proc.stderr, output_lines, "ERR"), daemon=True
    )
    t_out.start()
    t_err.start()

    with _JOBS_LOCK:
        _JOBS[job_id] = {
            "case_dir": str(target),
            "started_at": time.time(),
            "status": "running",
            "return_code": None,
            "output_lines": output_lines,
            "process": proc,
            "cmd": cmd,
        }

    # Reap the process and flip status when done (background thread).
    def _reaper() -> None:
        proc.wait()
        t_out.join(timeout=5)
        t_err.join(timeout=5)
        rc = proc.returncode
        with _JOBS_LOCK:
            if job_id in _JOBS:
                _JOBS[job_id]["status"] = "completed" if rc == 0 else "failed"
                _JOBS[job_id]["return_code"] = rc
                _JOBS[job_id]["process"] = None  # allow GC
            # Evict oldest finished jobs to cap memory usage.
            finished = [
                jid for jid, j in _JOBS.items()
                if j["status"] in ("completed", "failed", "cancelled")
            ]
            while len(finished) > _MAX_FINISHED_JOBS:
                oldest = min(finished, key=lambda jid: _JOBS[jid].get("started_at", 0))
                del _JOBS[oldest]
                finished.remove(oldest)

    threading.Thread(target=_reaper, daemon=True).start()

    return json.dumps({
        "job_id": job_id,
        "status": "running",
        "execution_mode": "docker" if use_docker else "native",
        "case_dir": str(target),
        "cmd": " ".join(cmd),
        "hint": "Poll with get_run_status(job_id='{job_id}') to follow progress.".format(job_id=job_id),
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 12 – get_run_status
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_run_status(
    job_id: str,
    tail: int = 60,
) -> str:
    """
    Check the status of a background analysis job started with run_analysis.

    Args:
        job_id: The job identifier returned by run_analysis.
        tail:   Number of most-recent output lines to include (default 60).

    Returns:
        JSON with status (running/completed/failed/cancelled), elapsed seconds,
        return_code, and the last *tail* lines of combined stdout/stderr.
    """
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if job is None:
            known = list(_JOBS.keys())
            return json.dumps({
                "error": f"Unknown job_id '{job_id}'.",
                "known_jobs": known,
            })

        lines = list(job["output_lines"])  # snapshot
        status = job["status"]
        rc = job["return_code"]
        elapsed = round(time.time() - job["started_at"], 1)
        case_dir = job["case_dir"]

    shown = lines[-tail:] if tail < len(lines) else lines
    return json.dumps({
        "job_id": job_id,
        "status": status,
        "elapsed_seconds": elapsed,
        "return_code": rc,
        "case_dir": case_dir,
        "output_lines_total": len(lines),
        "output_tail": shown,
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Tool 13 – cancel_run
# ─────────────────────────────────────────────────────────────────────────────

@mcp.tool()
def cancel_run(job_id: str) -> str:
    """
    Cancel a running analysis job.

    Sends SIGTERM to the process group so child tool processes (Hayabusa,
    Chainsaw, APT-Hunter) are also terminated.  If the process does not exit
    within 5 seconds a SIGKILL is sent.

    Args:
        job_id: The job identifier returned by run_analysis.
    """
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if job is None:
            return json.dumps({"error": f"Unknown job_id '{job_id}'."})
        proc = job.get("process")
        if proc is None:
            return json.dumps({
                "job_id": job_id,
                "status": job["status"],
                "message": "Job has already finished; nothing to cancel.",
            })

    try:
        os.killpg(os.getpgid(proc.pid), signal.SIGTERM)
    except (ProcessLookupError, PermissionError):
        try:
            proc.terminate()
        except Exception:
            pass

    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except Exception:
            proc.kill()
        proc.wait()

    with _JOBS_LOCK:
        if job_id in _JOBS:
            _JOBS[job_id]["status"] = "cancelled"
            _JOBS[job_id]["return_code"] = proc.returncode
            _JOBS[job_id]["process"] = None

    return json.dumps({
        "job_id": job_id,
        "status": "cancelled",
        "return_code": proc.returncode,
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    mcp.run()
